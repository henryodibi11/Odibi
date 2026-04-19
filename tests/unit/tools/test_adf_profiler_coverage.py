"""Tests for odibi.tools.adf_profiler — AdfProfiler REST API wrapper."""

from __future__ import annotations

import json
from unittest.mock import MagicMock, patch


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_profiler(**kwargs):
    """Create AdfProfiler with a mock credential (no real Azure auth)."""
    with patch("odibi.tools.adf_profiler.DefaultAzureCredential"):
        from odibi.tools.adf_profiler import AdfProfiler

        return AdfProfiler(
            subscription_id="sub-1",
            resource_group="rg-1",
            factory_name="factory-1",
            credential=MagicMock(),
            on_progress=lambda *a: None,
            **kwargs,
        )


def _mock_response(json_data, status=200):
    resp = MagicMock()
    resp.status_code = status
    resp.json.return_value = json_data
    resp.raise_for_status.return_value = None
    return resp


# =========================================================================
# Internal helpers
# =========================================================================


class TestInternalHelpers:
    def test_base_url(self):
        p = _make_profiler()
        url = p._base_url()
        assert "sub-1" in url
        assert "rg-1" in url
        assert "factory-1" in url

    def test_url_suffix(self):
        p = _make_profiler()
        url = p._url("pipelines")
        assert "pipelines" in url
        assert "api-version" in url

    def test_refresh_token(self):
        p = _make_profiler()
        mock_token = MagicMock()
        mock_token.token = "tok123"
        p.cred.get_token.return_value = mock_token
        p._refresh_token()
        assert p._headers["Authorization"] == "Bearer tok123"

    @patch("odibi.tools.adf_profiler.requests")
    def test_get(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response({"foo": 1})
        result = p._get("test")
        assert result == {"foo": 1}

    @patch("odibi.tools.adf_profiler.requests")
    def test_get_paged_single(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response({"value": [{"a": 1}]})
        result = p._get_paged("items")
        assert result == [{"a": 1}]

    @patch("odibi.tools.adf_profiler.requests")
    def test_get_paged_multiple_pages(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        page1 = _mock_response({"value": [{"a": 1}], "nextLink": "http://next"})
        page2 = _mock_response({"value": [{"b": 2}]})
        mock_requests.get.side_effect = [page1, page2]
        result = p._get_paged("items")
        assert len(result) == 2

    @patch("odibi.tools.adf_profiler.requests")
    def test_post(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response({"result": "ok"})
        result = p._post("action", {"key": "val"})
        assert result == {"result": "ok"}

    def test_cache_and_clear(self):
        p = _make_profiler()
        counter = {"calls": 0}

        def loader():
            counter["calls"] += 1
            return [1, 2, 3]

        r1 = p._cached("key1", loader)
        r2 = p._cached("key1", loader)
        assert r1 == r2
        assert counter["calls"] == 1
        p.clear_cache()
        p._cached("key1", loader)
        assert counter["calls"] == 2

    def test_default_progress(self):
        from odibi.tools.adf_profiler import AdfProfiler

        AdfProfiler._default_progress("step", 1, 5)  # just ensure no error

    def test_progress_callback(self):
        calls = []
        with patch("odibi.tools.adf_profiler.DefaultAzureCredential"):
            from odibi.tools.adf_profiler import AdfProfiler

            p = AdfProfiler(
                subscription_id="sub-1",
                resource_group="rg-1",
                factory_name="f1",
                credential=MagicMock(),
                on_progress=lambda s, c, t: calls.append((s, c, t)),
            )
        p._progress("test", 1, 3)
        assert calls == [("test", 1, 3)]


# =========================================================================
# Factory info
# =========================================================================


class TestFactoryInfo:
    @patch("odibi.tools.adf_profiler.requests")
    def test_get_factory_info(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "name": "factory-1",
                "location": "eastus",
                "tags": {"env": "dev"},
                "properties": {
                    "provisioningState": "Succeeded",
                    "createTime": "2024-01-01",
                    "version": "1.0",
                    "repoConfiguration": {
                        "type": "Git",
                        "repositoryName": "repo1",
                        "collaborationBranch": "main",
                    },
                    "globalParameters": {
                        "param1": {"type": "String", "value": "hello"},
                    },
                    "publicNetworkAccess": "Enabled",
                    "encryption": None,
                },
            }
        )
        info = p.get_factory_info()
        assert info["name"] == "factory-1"
        assert info["location"] == "eastus"
        assert info["provisioning_state"] == "Succeeded"
        assert info["global_parameters"]["param1"]["type"] == "String"
        assert info["repo_configuration"]["type"] == "Git"


# =========================================================================
# Pipelines
# =========================================================================


class TestPipelines:
    @patch("odibi.tools.adf_profiler.requests")
    def test_list_pipelines(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "pipe1",
                        "properties": {
                            "folder": {"name": "folder1"},
                            "description": "Test pipeline",
                            "concurrency": 1,
                            "parameters": {"p1": {"type": "String", "defaultValue": "x"}},
                            "variables": {"v1": {"type": "String", "defaultValue": "y"}},
                            "activities": [
                                {
                                    "name": "act1",
                                    "type": "Copy",
                                    "dependsOn": [],
                                    "typeProperties": {
                                        "source": {"type": "SqlSource"},
                                        "sink": {"type": "BlobSink"},
                                        "enableStaging": True,
                                    },
                                    "inputs": [{"referenceName": "ds_in", "parameters": {}}],
                                    "outputs": [{"referenceName": "ds_out", "parameters": {}}],
                                }
                            ],
                            "annotations": ["tag1"],
                        },
                    }
                ],
            }
        )
        pipes = p.list_pipelines()
        assert len(pipes) == 1
        assert pipes[0]["name"] == "pipe1"
        assert pipes[0]["activity_count"] == 1
        act = pipes[0]["activities"][0]
        assert act["source_type"] == "SqlSource"
        assert act["sink_type"] == "BlobSink"
        assert act["enable_staging"] is True
        assert act["inputs"][0]["dataset"] == "ds_in"

    def test_parse_activity_execute_pipeline(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "run_child",
                "type": "ExecutePipeline",
                "dependsOn": [{"activity": "act1", "dependencyConditions": ["Succeeded"]}],
                "typeProperties": {"pipeline": {"referenceName": "child_pipe"}},
            }
        )
        assert act["called_pipeline"] == "child_pipe"
        assert act["depends_on"][0]["activity"] == "act1"

    def test_parse_activity_if_condition(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "if1",
                "type": "IfCondition",
                "dependsOn": [],
                "typeProperties": {
                    "ifTrueActivities": [
                        {"name": "a1", "type": "Copy"},
                        {"name": "a2", "type": "Copy"},
                    ],
                    "ifFalseActivities": [
                        {"name": "a3", "type": "Wait"},
                    ],
                },
            }
        )
        assert act["sub_activity_count"] == 3

    def test_parse_activity_switch(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "switch1",
                "type": "Switch",
                "dependsOn": [],
                "typeProperties": {
                    "cases": [
                        {
                            "activities": [
                                {"name": "c1", "type": "Copy"},
                                {"name": "c2", "type": "Wait"},
                            ]
                        },
                    ],
                    "defaultActivities": [{"name": "d1", "type": "Wait"}],
                },
            }
        )
        assert act["sub_activity_count"] == 3

    def test_parse_activity_foreach(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "fe1",
                "type": "ForEach",
                "dependsOn": [],
                "typeProperties": {
                    "activities": [{"name": "a1", "type": "Copy"}],
                },
            }
        )
        assert act["sub_activity_count"] == 1

    def test_parse_activity_web(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "web1",
                "type": "WebActivity",
                "dependsOn": [],
                "typeProperties": {"method": "GET", "url": "https://example.com"},
            }
        )
        assert act["method"] == "GET"
        assert act["url"] == "https://example.com"

    def test_parse_activity_lookup(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "lk1",
                "type": "Lookup",
                "dependsOn": [],
                "typeProperties": {"source": {"type": "SqlSource"}},
            }
        )
        assert act["source_type"] == "SqlSource"

    def test_parse_activity_get_metadata(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "gm1",
                "type": "GetMetadata",
                "dependsOn": [],
                "typeProperties": {"source": {"type": "BlobSource"}},
            }
        )
        assert act["source_type"] == "BlobSource"

    def test_parse_activity_execute_data_flow(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "edf1",
                "type": "ExecuteDataFlow",
                "dependsOn": [],
                "typeProperties": {"dataFlow": {"referenceName": "df_transform"}},
            }
        )
        assert act["data_flow"] == "df_transform"

    def test_parse_activity_no_io(self):
        p = _make_profiler()
        act = p._parse_activity(
            {
                "name": "wait1",
                "type": "Wait",
                "dependsOn": [],
                "typeProperties": {},
            }
        )
        assert "inputs" not in act
        assert "outputs" not in act

    def test_count_nested_empty(self):
        p = _make_profiler()
        assert p._count_nested_activities({}) == 0

    def test_count_nested_cases_with_activities(self):
        p = _make_profiler()
        count = p._count_nested_activities(
            {
                "cases": [
                    {"activities": [{"type": "Copy"}, {"type": "Wait"}]},
                    {"activities": [{"type": "Copy"}]},
                ],
            }
        )
        assert count == 3


# =========================================================================
# Datasets
# =========================================================================


class TestDatasets:
    @patch("odibi.tools.adf_profiler.requests")
    def test_list_datasets(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "ds1",
                        "properties": {
                            "type": "DelimitedText",
                            "folder": {"name": "raw"},
                            "linkedServiceName": {"referenceName": "ls_blob"},
                            "description": "CSV data",
                            "parameters": {"file": {"type": "String", "defaultValue": "test.csv"}},
                            "schema": [{"name": "col1", "type": "String"}],
                            "structure": [],
                            "annotations": [],
                            "typeProperties": {
                                "location": {
                                    "type": "AzureBlobStorageLocation",
                                    "container": "raw",
                                    "folderPath": "data",
                                    "fileName": "test.csv",
                                },
                            },
                        },
                    }
                ],
            }
        )
        datasets = p.list_datasets()
        assert len(datasets) == 1
        assert datasets[0]["name"] == "ds1"
        assert datasets[0]["linked_service"] == "ls_blob"
        assert datasets[0]["location"]["container"] == "raw"

    def test_extract_location_with_location_key(self):
        p = _make_profiler()
        loc = p._extract_location(
            {
                "location": {
                    "type": "AzureBlobStorageLocation",
                    "container": "cont1",
                    "fileSystem": "fs1",
                    "folderPath": "path/to",
                    "fileName": "file.csv",
                },
            }
        )
        assert loc["type"] == "AzureBlobStorageLocation"
        assert loc["container"] == "cont1"

    def test_extract_location_without_location_key(self):
        p = _make_profiler()
        loc = p._extract_location(
            {
                "tableName": "dbo.table1",
                "schema": "dbo",
            }
        )
        assert loc["tableName"] == "dbo.table1"
        assert loc["schema"] == "dbo"

    def test_extract_location_empty(self):
        p = _make_profiler()
        loc = p._extract_location({})
        assert loc == {}

    def test_extract_location_filesystem_fallback(self):
        p = _make_profiler()
        loc = p._extract_location(
            {
                "location": {
                    "type": "AzureBlobFSLocation",
                    "fileSystem": "myfs",
                },
            }
        )
        assert loc["container"] == "myfs"


# =========================================================================
# Linked Services
# =========================================================================


class TestLinkedServices:
    @patch("odibi.tools.adf_profiler.requests")
    def test_list_linked_services(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "ls1",
                        "properties": {
                            "type": "AzureBlobStorage",
                            "description": "Blob storage",
                            "annotations": [],
                            "connectVia": {"referenceName": "ir-default"},
                            "typeProperties": {
                                "connectionString": "DefaultEndpoints...",
                                "accountName": "myaccount",
                            },
                        },
                    }
                ],
            }
        )
        services = p.list_linked_services()
        assert len(services) == 1
        assert services[0]["name"] == "ls1"
        assert services[0]["connect_via"] == "ir-default"

    def test_safe_connection_info_basic(self):
        p = _make_profiler()
        info = p._safe_connection_info(
            {
                "url": "https://example.com",
                "server": "myserver",
                "database": "mydb",
                "authenticationType": "SQL",
            }
        )
        assert info["url"] == "https://example.com"
        assert info["server"] == "myserver"

    def test_safe_connection_info_dict_value(self):
        p = _make_profiler()
        info = p._safe_connection_info(
            {
                "connectionString": {"value": "Server=x;Database=y"},
            }
        )
        assert info["connectionString"] == "Server=x;Database=y"

    def test_safe_connection_info_ignores_unknown_keys(self):
        p = _make_profiler()
        info = p._safe_connection_info(
            {
                "unknownKey": "should_not_appear",
                "url": "https://test",
            }
        )
        assert "unknownKey" not in info
        assert info["url"] == "https://test"


# =========================================================================
# Triggers
# =========================================================================


class TestTriggers:
    @patch("odibi.tools.adf_profiler.requests")
    def test_schedule_trigger(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "trig1",
                        "properties": {
                            "type": "ScheduleTrigger",
                            "runtimeState": "Started",
                            "description": "Daily trigger",
                            "annotations": [],
                            "typeProperties": {
                                "recurrence": {
                                    "frequency": "Day",
                                    "interval": 1,
                                    "startTime": "2024-01-01T06:00:00Z",
                                    "timeZone": "UTC",
                                },
                            },
                            "pipelines": [
                                {"pipelineReference": {"referenceName": "pipe1"}, "parameters": {}},
                            ],
                        },
                    }
                ],
            }
        )
        triggers = p.list_triggers()
        assert len(triggers) == 1
        assert triggers[0]["schedule"]["frequency"] == "Day"
        assert triggers[0]["associated_pipelines"][0]["pipeline"] == "pipe1"

    @patch("odibi.tools.adf_profiler.requests")
    def test_tumbling_window_trigger(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "tw1",
                        "properties": {
                            "type": "TumblingWindowTrigger",
                            "runtimeState": "Started",
                            "typeProperties": {
                                "frequency": "Hour",
                                "interval": 1,
                                "startTime": "2024-01-01",
                                "maxConcurrency": 5,
                                "retryPolicy": {"count": 3},
                                "pipeline": {
                                    "pipelineReference": {"referenceName": "pipe2"},
                                    "parameters": {"p": "v"},
                                },
                            },
                        },
                    }
                ],
            }
        )
        triggers = p.list_triggers()
        assert triggers[0]["tumbling_window"]["frequency"] == "Hour"
        assert triggers[0]["associated_pipelines"][0]["pipeline"] == "pipe2"

    @patch("odibi.tools.adf_profiler.requests")
    def test_blob_events_trigger(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "be1",
                        "properties": {
                            "type": "BlobEventsTrigger",
                            "runtimeState": "Stopped",
                            "typeProperties": {
                                "blobPathBeginsWith": "/container/path",
                                "blobPathEndsWith": ".csv",
                                "events": ["Microsoft.Storage.BlobCreated"],
                                "scope": "/subscriptions/sub/...",
                            },
                            "pipelines": [
                                {"pipelineReference": {"referenceName": "pipe3"}, "parameters": {}},
                            ],
                        },
                    }
                ],
            }
        )
        triggers = p.list_triggers()
        assert triggers[0]["blob_events"]["events"] == ["Microsoft.Storage.BlobCreated"]

    @patch("odibi.tools.adf_profiler.requests")
    def test_trigger_no_pipelines(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "t_none",
                        "properties": {
                            "type": "CustomEventsTrigger",
                            "runtimeState": "Started",
                            "typeProperties": {},
                        },
                    }
                ],
            }
        )
        triggers = p.list_triggers()
        assert "associated_pipelines" not in triggers[0]


# =========================================================================
# Data Flows
# =========================================================================


class TestDataFlows:
    @patch("odibi.tools.adf_profiler.requests")
    def test_list_data_flows(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "df1",
                        "properties": {
                            "type": "MappingDataFlow",
                            "folder": {"name": "transforms"},
                            "description": "Transform flow",
                            "typeProperties": {
                                "sources": [
                                    {
                                        "name": "src1",
                                        "dataset": {"referenceName": "ds_in"},
                                        "linkedService": {},
                                    },
                                ],
                                "sinks": [
                                    {
                                        "name": "sink1",
                                        "dataset": {"referenceName": "ds_out"},
                                        "linkedService": {},
                                    },
                                ],
                                "transformations": [
                                    {"name": "derive1", "description": "Add column"},
                                ],
                                "script": "source(output(col1 as string))",
                                "scriptLines": ["line1", "line2"],
                            },
                        },
                    }
                ],
            }
        )
        flows = p.list_data_flows()
        assert len(flows) == 1
        assert flows[0]["name"] == "df1"
        assert flows[0]["sources"][0]["dataset"] == "ds_in"
        assert flows[0]["sinks"][0]["dataset"] == "ds_out"
        assert flows[0]["transformations"][0]["name"] == "derive1"
        assert flows[0]["script"] is not None


# =========================================================================
# Integration Runtimes
# =========================================================================


class TestIntegrationRuntimes:
    @patch("odibi.tools.adf_profiler.requests")
    def test_managed_ir(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        # First call: list, second call: getStatus
        list_resp = _mock_response(
            {
                "value": [
                    {
                        "name": "ir-managed",
                        "properties": {
                            "type": "Managed",
                            "description": "Managed IR",
                            "typeProperties": {
                                "computeProperties": {
                                    "location": "AutoResolve",
                                    "dataFlowProperties": {
                                        "coreCount": 8,
                                        "computeType": "General",
                                        "timeToLive": 10,
                                    },
                                },
                            },
                        },
                    }
                ],
            }
        )
        status_resp = _mock_response({"properties": {"state": "Online"}})
        mock_requests.get.return_value = list_resp
        mock_requests.post.return_value = status_resp
        runtimes = p.list_integration_runtimes()
        assert len(runtimes) == 1
        assert runtimes[0]["compute"]["core_count"] == 8
        assert runtimes[0]["state"] == "Online"

    @patch("odibi.tools.adf_profiler.requests")
    def test_self_hosted_ir(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "ir-sh",
                        "properties": {
                            "type": "SelfHosted",
                            "description": "Self-hosted IR",
                            "typeProperties": {},
                        },
                    }
                ],
            }
        )
        mock_requests.post.return_value = _mock_response({"properties": {"state": "Online"}})
        runtimes = p.list_integration_runtimes()
        assert runtimes[0]["self_hosted"] is True

    @patch("odibi.tools.adf_profiler.requests")
    def test_ir_status_error(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response(
            {
                "value": [
                    {
                        "name": "ir-err",
                        "properties": {"type": "Managed", "typeProperties": {}},
                    }
                ],
            }
        )
        mock_requests.post.side_effect = Exception("status failed")
        runtimes = p.list_integration_runtimes()
        assert runtimes[0]["state"] == "Unknown"


# =========================================================================
# Pipeline Runs & Activity Runs
# =========================================================================


class TestPipelineRuns:
    @patch("odibi.tools.adf_profiler.requests")
    def test_get_pipeline_runs_basic(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response(
            {
                "value": [
                    {
                        "runId": "r1",
                        "pipelineName": "pipe1",
                        "status": "Succeeded",
                        "runStart": "2024-01-01T00:00:00Z",
                        "runEnd": "2024-01-01T00:05:00Z",
                        "durationInMs": 300000,
                        "invokedBy": {"name": "Manual", "invokedByType": "Manual"},
                        "parameters": {"p": "v"},
                        "message": "",
                        "runGroupId": "rg1",
                    }
                ],
            }
        )
        runs = p.get_pipeline_runs(days_back=7)
        assert len(runs) == 1
        assert runs[0]["run_id"] == "r1"
        assert runs[0]["duration_ms"] == 300000

    @patch("odibi.tools.adf_profiler.requests")
    def test_get_pipeline_runs_with_filters(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response({"value": []})
        runs = p.get_pipeline_runs(pipeline_name="p1", status="Failed")
        assert runs == []
        # Verify filters were passed
        call_body = mock_requests.post.call_args[1]["json"]
        assert len(call_body["filters"]) == 2

    @patch("odibi.tools.adf_profiler.requests")
    def test_get_activity_runs(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response(
            {
                "value": [
                    {
                        "activityName": "Copy1",
                        "activityType": "Copy",
                        "status": "Succeeded",
                        "activityRunStart": "2024-01-01T00:00:00Z",
                        "activityRunEnd": "2024-01-01T00:01:00Z",
                        "durationInMs": 60000,
                        "error": None,
                        "input": {},
                        "output": {},
                    }
                ],
            }
        )
        runs = p.get_activity_runs("r1", "pipe1")
        assert len(runs) == 1
        assert runs[0]["activity_name"] == "Copy1"

    @patch("odibi.tools.adf_profiler.requests")
    def test_get_last_run_exists(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response(
            {
                "value": [
                    {
                        "runId": "latest",
                        "pipelineName": "p1",
                        "status": "Succeeded",
                        "runStart": None,
                        "runEnd": None,
                        "durationInMs": 100,
                        "invokedBy": {},
                        "parameters": {},
                        "message": None,
                        "runGroupId": None,
                    }
                ],
            }
        )
        run = p.get_last_run("p1")
        assert run["run_id"] == "latest"

    @patch("odibi.tools.adf_profiler.requests")
    def test_get_last_run_none(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response({"value": []})
        run = p.get_last_run("p1")
        assert run is None


# =========================================================================
# Pipeline Run Stats
# =========================================================================


class TestPipelineRunStats:
    @patch("odibi.tools.adf_profiler.requests")
    def test_stats_no_runs(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response({"value": []})
        stats = p.get_pipeline_run_stats("p1")
        assert stats == {"total_runs": 0}

    @patch("odibi.tools.adf_profiler.requests")
    def test_stats_with_runs(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response(
            {
                "value": [
                    {
                        "runId": "r1",
                        "pipelineName": "p1",
                        "status": "Succeeded",
                        "runStart": "2024-01-01",
                        "runEnd": "2024-01-01",
                        "durationInMs": 10000,
                        "invokedBy": {},
                        "parameters": {},
                        "message": None,
                        "runGroupId": None,
                    },
                    {
                        "runId": "r2",
                        "pipelineName": "p1",
                        "status": "Failed",
                        "runStart": "2024-01-02",
                        "runEnd": "2024-01-02",
                        "durationInMs": 5000,
                        "invokedBy": {},
                        "parameters": {},
                        "message": None,
                        "runGroupId": None,
                    },
                ],
            }
        )
        stats = p.get_pipeline_run_stats("p1")
        assert stats["total_runs"] == 2
        assert stats["success_rate"] == 50.0
        assert stats["avg_duration_sec"] == 7.5
        assert stats["min_duration_sec"] == 5.0
        assert stats["max_duration_sec"] == 10.0

    @patch("odibi.tools.adf_profiler.requests")
    def test_stats_no_durations(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.post.return_value = _mock_response(
            {
                "value": [
                    {
                        "runId": "r1",
                        "pipelineName": "p1",
                        "status": "InProgress",
                        "runStart": "2024-01-01",
                        "runEnd": None,
                        "durationInMs": None,
                        "invokedBy": {},
                        "parameters": {},
                        "message": None,
                        "runGroupId": None,
                    },
                ],
            }
        )
        stats = p.get_pipeline_run_stats("p1")
        assert stats["avg_duration_sec"] is None


# =========================================================================
# Lineage
# =========================================================================


class TestLineage:
    def test_get_pipeline_lineage_found(self):
        p = _make_profiler()
        p._cache["pipelines"] = [
            {
                "name": "pipe1",
                "activities": [
                    {
                        "name": "copy1",
                        "type": "Copy",
                        "inputs": [{"dataset": "ds_in"}],
                        "outputs": [{"dataset": "ds_out"}],
                    }
                ],
            }
        ]
        p._cache["datasets"] = [
            {"name": "ds_in", "type": "DelimitedText", "linked_service": "ls_blob", "location": {}},
            {"name": "ds_out", "type": "Parquet", "linked_service": "ls_adls", "location": {}},
        ]
        p._cache["linked_services"] = [
            {"name": "ls_blob", "type": "AzureBlobStorage"},
            {"name": "ls_adls", "type": "AzureBlobFS"},
        ]
        lineage = p.get_pipeline_lineage("pipe1")
        assert lineage["pipeline"] == "pipe1"
        assert len(lineage["lineage"]) == 1
        assert lineage["lineage"][0]["sources"][0]["dataset"] == "ds_in"
        assert lineage["lineage"][0]["sinks"][0]["dataset"] == "ds_out"

    def test_get_pipeline_lineage_not_found(self):
        p = _make_profiler()
        p._cache["pipelines"] = []
        p._cache["datasets"] = []
        p._cache["linked_services"] = []
        lineage = p.get_pipeline_lineage("missing_pipe")
        assert "error" in lineage

    def test_get_pipeline_lineage_no_io_activity(self):
        p = _make_profiler()
        p._cache["pipelines"] = [
            {
                "name": "pipe1",
                "activities": [{"name": "wait1", "type": "Wait"}],
            }
        ]
        p._cache["datasets"] = []
        p._cache["linked_services"] = []
        lineage = p.get_pipeline_lineage("pipe1")
        assert len(lineage["lineage"]) == 0


# =========================================================================
# Full Profile
# =========================================================================


class TestFullProfile:
    def test_full_profile_no_runs(self):
        p = _make_profiler()
        p.get_factory_info = MagicMock(
            return_value={
                "name": "f1",
                "location": "eastus",
                "provisioning_state": "Succeeded",
                "create_time": "2024",
                "version": "1",
                "repo_configuration": None,
                "global_parameters": {},
                "public_network_access": "Enabled",
                "encryption": None,
                "tags": {},
            }
        )
        p.list_pipelines = MagicMock(return_value=[])
        p.list_datasets = MagicMock(return_value=[])
        p.list_linked_services = MagicMock(return_value=[])
        p.list_triggers = MagicMock(return_value=[])
        p.list_data_flows = MagicMock(return_value=[])
        p.list_integration_runtimes = MagicMock(return_value=[])

        profile = p.full_profile(include_runs=False)
        assert profile["factory"]["name"] == "f1"
        assert profile["summary"]["pipeline_count"] == 0
        assert "run_statistics" not in profile
        assert "profile_duration_sec" in profile

    def test_full_profile_with_runs(self):
        p = _make_profiler()
        p.get_factory_info = MagicMock(
            return_value={
                "name": "f1",
                "location": "eastus",
                "provisioning_state": "Succeeded",
                "create_time": "2024",
                "version": "1",
                "repo_configuration": None,
                "global_parameters": {},
                "public_network_access": "Enabled",
                "encryption": None,
                "tags": {},
            }
        )
        p.list_pipelines = MagicMock(
            return_value=[{"name": "p1", "activities": [{"type": "Copy"}]}]
        )
        p.list_datasets = MagicMock(return_value=[{"type": "Parquet"}])
        p.list_linked_services = MagicMock(return_value=[{"type": "AzureBlobFS"}])
        p.list_triggers = MagicMock(return_value=[])
        p.list_data_flows = MagicMock(return_value=[])
        p.list_integration_runtimes = MagicMock(return_value=[])
        p._parallel_run_stats = MagicMock(return_value={"p1": {"total_runs": 5}})

        profile = p.full_profile(include_runs=True, days_back=7)
        assert "run_statistics" in profile
        assert profile["run_statistics"]["p1"]["total_runs"] == 5
        assert profile["summary"]["pipeline_count"] == 1


# =========================================================================
# Parallel Run Stats
# =========================================================================


class TestParallelRunStats:
    def test_parallel_run_stats_success(self):
        p = _make_profiler(max_workers=2)
        p.get_pipeline_run_stats = MagicMock(return_value={"total_runs": 10})
        pipelines = [{"name": "p1"}, {"name": "p2"}]
        stats = p._parallel_run_stats(pipelines, days_back=7)
        assert "p1" in stats
        assert "p2" in stats
        assert stats["p1"]["total_runs"] == 10

    def test_parallel_run_stats_error(self):
        p = _make_profiler(max_workers=1)
        p.get_pipeline_run_stats = MagicMock(side_effect=Exception("API error"))
        pipelines = [{"name": "p1"}]
        stats = p._parallel_run_stats(pipelines, days_back=7)
        assert "error" in stats["p1"]


# =========================================================================
# Report Generation
# =========================================================================


class TestReportGeneration:
    def _make_profile(self):
        return {
            "profiled_at": "2024-01-01T00:00:00Z",
            "factory": {
                "name": "f1",
                "location": "eastus",
                "provisioning_state": "Succeeded",
                "create_time": "2024-01-01",
                "version": "1",
                "repo_configuration": {
                    "type": "Git",
                    "repositoryName": "repo",
                    "collaborationBranch": "main",
                },
                "global_parameters": {"gp1": {"type": "String", "value": "val1"}},
            },
            "summary": {
                "pipeline_count": 1,
                "dataset_count": 1,
                "linked_service_count": 1,
                "trigger_count": 1,
                "data_flow_count": 1,
                "integration_runtime_count": 1,
                "connector_types": ["AzureBlobFS"],
                "dataset_types": ["Parquet"],
                "activity_types": ["Copy"],
            },
            "pipelines": [
                {
                    "name": "pipe1",
                    "folder": "folder1",
                    "description": "Test pipeline",
                    "parameters": {"p1": {"type": "String", "default": "x"}},
                    "activities": [
                        {"name": "a1", "type": "Copy", "depends_on": [{"activity": "a0"}]}
                    ],
                }
            ],
            "datasets": [
                {"name": "ds1", "type": "Parquet", "linked_service": "ls1", "folder": "raw"}
            ],
            "linked_services": [
                {
                    "name": "ls1",
                    "type": "AzureBlobFS",
                    "connect_via": "ir1",
                    "connection_details": {"authenticationType": "MSI"},
                }
            ],
            "triggers": [
                {
                    "name": "t1",
                    "type": "ScheduleTrigger",
                    "runtime_state": "Started",
                    "associated_pipelines": [{"pipeline": "pipe1"}],
                }
            ],
            "data_flows": [
                {
                    "name": "df1",
                    "sources": [{"name": "src1", "dataset": "ds_in"}],
                    "sinks": [{"name": "sink1", "dataset": "ds_out"}],
                    "transformations": [{"name": "t1"}],
                }
            ],
            "integration_runtimes": [{"name": "ir1", "type": "Managed", "state": "Online"}],
            "run_statistics": {
                "pipe1": {
                    "total_runs": 5,
                    "success_rate": 80.0,
                    "avg_duration_sec": 10.0,
                    "last_run": {"status": "Succeeded", "run_end": "2024-01-01"},
                },
                "pipe2": {"error": "API failed"},
                "pipe3": {"total_runs": 0},
            },
        }

    def test_build_report_lines(self):
        p = _make_profiler()
        lines = p._build_report_lines(self._make_profile())
        report = "\n".join(lines)
        assert "# Azure Data Factory Profile: f1" in report
        assert "## Factory Overview" in report
        assert "## Resource Summary" in report
        assert "## Linked Services" in report
        assert "## Datasets" in report
        assert "## Pipelines" in report
        assert "### pipe1" in report
        assert "## Triggers" in report
        assert "## Data Flows" in report
        assert "## Integration Runtimes" in report
        assert "## Pipeline Run Statistics" in report
        assert "80.0%" in report
        assert "API failed" in report
        assert "Never" in report  # pipe3 with 0 runs
        assert "Git" in report  # repo type

    def test_build_report_lines_no_data_flows(self):
        p = _make_profiler()
        profile = self._make_profile()
        profile["data_flows"] = []
        del profile["run_statistics"]
        lines = p._build_report_lines(profile)
        report = "\n".join(lines)
        assert "## Data Flows" not in report

    def test_generate_report_no_file(self):
        p = _make_profiler()
        p.full_profile = MagicMock(return_value=self._make_profile())
        report = p.generate_report(output_path=None)
        assert "# Azure Data Factory Profile" in report

    def test_generate_report_to_file(self, tmp_path):
        p = _make_profiler()
        p.full_profile = MagicMock(return_value=self._make_profile())
        out = tmp_path / "report.md"
        report = p.generate_report(output_path=str(out))
        assert out.read_text(encoding="utf-8") == report

    def test_print_summary(self, capsys):
        p = _make_profiler()
        p.full_profile = MagicMock(
            return_value={
                "factory": {"name": "f1", "location": "eastus"},
                "summary": {
                    "pipeline_count": 3,
                    "dataset_count": 5,
                    "linked_service_count": 2,
                    "trigger_count": 1,
                    "data_flow_count": 0,
                    "integration_runtime_count": 1,
                    "connector_types": ["Blob"],
                    "dataset_types": ["CSV"],
                },
            }
        )
        p.print_summary()
        captured = capsys.readouterr()
        assert "f1" in captured.out
        assert "Pipelines" in captured.out

    def test_to_json_no_file(self):
        p = _make_profiler()
        p.full_profile = MagicMock(return_value={"factory": "f1"})
        result = p.to_json(output_path=None)
        assert json.loads(result)["factory"] == "f1"

    def test_to_json_to_file(self, tmp_path):
        p = _make_profiler()
        p.full_profile = MagicMock(return_value={"factory": "f1"})
        out = tmp_path / "profile.json"
        result = p.to_json(output_path=str(out))
        assert out.read_text(encoding="utf-8") == result


# =========================================================================
# Excel Report
# =========================================================================


class TestExcelReport:
    def _make_full_profile(self):
        return {
            "profiled_at": "2024-01-01T00:00:00Z",
            "profile_duration_sec": 5.0,
            "factory": {
                "name": "f1",
                "location": "eastus",
                "provisioning_state": "Succeeded",
                "create_time": "2024-01-01",
                "version": "1",
                "repo_configuration": {
                    "type": "Git",
                    "repositoryName": "repo",
                    "collaborationBranch": "main",
                },
                "global_parameters": {"gp1": {"type": "String", "value": "val"}},
            },
            "summary": {
                "pipeline_count": 1,
                "dataset_count": 1,
                "linked_service_count": 1,
                "trigger_count": 1,
                "data_flow_count": 1,
                "integration_runtime_count": 1,
                "connector_types": ["Blob"],
                "dataset_types": ["CSV"],
                "activity_types": ["Copy"],
            },
            "pipelines": [
                {
                    "name": "p1",
                    "folder": "f",
                    "description": "d",
                    "activity_count": 1,
                    "concurrency": 1,
                    "parameters": {"x": {"type": "String"}},
                    "variables": {"v": {"type": "String"}},
                    "activities": [
                        {
                            "name": "a1",
                            "type": "Copy",
                            "depends_on": [{"activity": "a0"}],
                            "inputs": [{"dataset": "ds_in"}],
                            "outputs": [{"dataset": "ds_out"}],
                            "source_type": "Sql",
                            "sink_type": "Blob",
                        }
                    ],
                }
            ],
            "datasets": [
                {
                    "name": "ds1",
                    "type": "CSV",
                    "linked_service": "ls1",
                    "folder": "raw",
                    "location": {"type": "Blob", "container": "c"},
                    "parameters": {},
                    "description": "d",
                }
            ],
            "linked_services": [
                {
                    "name": "ls1",
                    "type": "Blob",
                    "connect_via": "ir1",
                    "connection_details": {"authenticationType": "Key", "accountName": "acc1"},
                    "description": "d",
                }
            ],
            "triggers": [
                {
                    "name": "t1",
                    "type": "ScheduleTrigger",
                    "runtime_state": "Started",
                    "schedule": {"interval": 1, "frequency": "Day", "time_zone": "UTC"},
                    "associated_pipelines": [{"pipeline": "p1"}],
                    "description": "d",
                }
            ],
            "data_flows": [
                {
                    "name": "df1",
                    "type": "MappingDataFlow",
                    "folder": "f",
                    "sources": [{"name": "s", "dataset": "ds_in"}],
                    "sinks": [{"name": "sk", "dataset": "ds_out"}],
                    "transformations": [{"name": "t"}],
                    "description": "d",
                }
            ],
            "integration_runtimes": [
                {
                    "name": "ir1",
                    "type": "Managed",
                    "state": "Online",
                    "compute": {
                        "location": "AutoResolve",
                        "core_count": 8,
                        "compute_type": "General",
                    },
                    "description": "d",
                }
            ],
            "run_statistics": {
                "p1": {
                    "total_runs": 5,
                    "success_rate": 80.0,
                    "avg_duration_sec": 10,
                    "min_duration_sec": 5,
                    "max_duration_sec": 15,
                    "last_run": {
                        "status": "Succeeded",
                        "run_start": "2024",
                        "run_end": "2024",
                        "triggered_by": "Manual",
                    },
                },
                "p2": {"error": "failed"},
            },
        }

    def test_to_excel(self, tmp_path):
        p = _make_profiler()
        p.full_profile = MagicMock(return_value=self._make_full_profile())
        out = tmp_path / "report.xlsx"
        result = p.to_excel(output_path=str(out))
        assert result == str(out)
        assert out.exists()
        # Verify sheets
        import openpyxl

        wb = openpyxl.load_workbook(str(out))
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Pipelines" in sheet_names
        assert "Activities" in sheet_names
        assert "Datasets" in sheet_names
        assert "Linked Services" in sheet_names
        assert "Triggers" in sheet_names
        assert "Data Flows" in sheet_names
        assert "Integration Runtimes" in sheet_names
        assert "Run Statistics" in sheet_names
        assert "Lineage" in sheet_names
        wb.close()

    def test_to_excel_no_data_flows(self, tmp_path):
        p = _make_profiler()
        profile = self._make_full_profile()
        profile["data_flows"] = []
        p.full_profile = MagicMock(return_value=profile)
        out = tmp_path / "report_nodf.xlsx"
        p.to_excel(output_path=str(out))
        assert out.exists()
        import openpyxl

        wb = openpyxl.load_workbook(str(out))
        assert "Data Flows" not in wb.sheetnames
        wb.close()

    def test_to_excel_tumbling_window_trigger(self, tmp_path):
        p = _make_profiler()
        profile = self._make_full_profile()
        profile["triggers"] = [
            {
                "name": "tw1",
                "type": "TumblingWindowTrigger",
                "runtime_state": "Started",
                "tumbling_window": {"interval": 1, "frequency": "Hour"},
                "associated_pipelines": [],
                "description": "",
            }
        ]
        p.full_profile = MagicMock(return_value=profile)
        out = tmp_path / "report_tw.xlsx"
        p.to_excel(output_path=str(out))
        assert out.exists()

    def test_to_excel_blob_events_trigger(self, tmp_path):
        p = _make_profiler()
        profile = self._make_full_profile()
        profile["triggers"] = [
            {
                "name": "be1",
                "type": "BlobEventsTrigger",
                "runtime_state": "Started",
                "blob_events": {"events": ["BlobCreated", "BlobDeleted"]},
                "associated_pipelines": [],
                "description": "",
            }
        ]
        p.full_profile = MagicMock(return_value=profile)
        out = tmp_path / "report_be.xlsx"
        p.to_excel(output_path=str(out))
        assert out.exists()

    def test_to_excel_no_lineage_rows(self, tmp_path):
        p = _make_profiler()
        profile = self._make_full_profile()
        # Remove inputs/outputs from activities
        for pipe in profile["pipelines"]:
            for act in pipe["activities"]:
                act.pop("inputs", None)
                act.pop("outputs", None)
        p.full_profile = MagicMock(return_value=profile)
        out = tmp_path / "report_nolin.xlsx"
        p.to_excel(output_path=str(out))
        import openpyxl

        wb = openpyxl.load_workbook(str(out))
        assert "Lineage" not in wb.sheetnames
        wb.close()

    def test_auto_fit_columns(self):
        p = _make_profiler()
        mock_writer = MagicMock()
        mock_cell = MagicMock()
        mock_cell.value = "test value"
        mock_cell.column_letter = "A"
        mock_col = [mock_cell]
        mock_ws = MagicMock()
        mock_ws.columns = [mock_col]
        mock_writer.sheets = {"Sheet1": mock_ws}
        p._auto_fit_columns(mock_writer)
        mock_ws.column_dimensions.__getitem__.assert_called()


# =========================================================================
# Caching behavior
# =========================================================================


class TestCaching:
    @patch("odibi.tools.adf_profiler.requests")
    def test_list_methods_use_cache(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response({"value": []})
        # Call twice — only one API call
        p.list_pipelines()
        p.list_pipelines()
        assert mock_requests.get.call_count == 1

    @patch("odibi.tools.adf_profiler.requests")
    def test_cache_cleared(self, mock_requests):
        p = _make_profiler()
        p.cred.get_token.return_value = MagicMock(token="t")
        mock_requests.get.return_value = _mock_response({"value": []})
        p.list_datasets()
        p.clear_cache()
        p.list_datasets()
        assert mock_requests.get.call_count == 2
