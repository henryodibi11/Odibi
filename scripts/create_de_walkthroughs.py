"""
Generate professional-grade Data Engineering Planning Walkthroughs.
Creates Bronze, Silver, and Gold layer workbooks with step-by-step guidance.
Enhanced with Odibi-specific patterns, validation contracts, FK relationships, and semantic layer.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# =============================================================================
# STYLING CONSTANTS
# =============================================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
PROMPT_FONT = Font(italic=True, color="666666", size=10)
WARNING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FONT = Font(color="9C0006", bold=True)
TIP_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
TIP_FONT = Font(color="006100")
ENCOURAGE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ENCOURAGE_FONT = Font(color="806000", italic=True)
EXAMPLE_FONT = Font(italic=True, color="808080", size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def add_title(ws, text, row=1):
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")


def add_prompt(ws, text, row, col=1):
    cell = ws.cell(row=row, column=col, value=f"→ {text}")
    cell.font = PROMPT_FONT


def add_warning(ws, text, row, col=1, span=3):
    cell = ws.cell(row=row, column=col, value=f"⚠️ {text}")
    cell.fill = WARNING_FILL
    cell.font = WARNING_FONT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)


def add_tip(ws, text, row, col=1, span=3):
    cell = ws.cell(row=row, column=col, value=f"✅ {text}")
    cell.fill = TIP_FILL
    cell.font = TIP_FONT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)


def add_encouragement(ws, text, row, col=1, span=3):
    cell = ws.cell(row=row, column=col, value=f"💡 {text}")
    cell.fill = ENCOURAGE_FILL
    cell.font = ENCOURAGE_FONT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)


def add_data_rows(ws, start_row, num_rows, num_cols):
    for row in range(start_row, start_row + num_rows):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


# =============================================================================
# BRONZE WALKTHROUGH
# =============================================================================


def create_bronze_walkthrough(output_dir):
    wb = Workbook()

    # B0 - Overview
    ws = wb.active
    ws.title = "B0_Overview"

    add_title(ws, "BRONZE LAYER PLANNING WALKTHROUGH")

    ws.cell(row=3, column=1, value="What is Bronze?")
    ws.cell(row=3, column=1).font = SECTION_FONT
    ws.cell(
        row=4,
        column=1,
        value='"Land it as-is. No cleaning, no joins, no business logic. Bronze is your undo button."',
    )
    ws.cell(row=4, column=1).font = Font(italic=True, size=11)

    ws.cell(row=6, column=1, value="How to Use This Workbook")
    ws.cell(row=6, column=1).font = SECTION_FONT

    steps = [
        "1. Fill B1_Source_Inventory - What source systems and tables exist?",
        "2. Fill B2_Node_Design - One row per Bronze node (scalable table format)",
        "3. Fill B3_Column_Mapping - Map source columns to Bronze columns",
        "4. Plan validations in B4_Validation - What checks ensure quality?",
        "5. For migrations, use B5_Migration - Map old to new",
    ]
    for i, step in enumerate(steps, 7):
        ws.cell(row=i, column=1, value=step)

    add_encouragement(
        ws, "Take your time here. 30 minutes of planning saves 3 hours of debugging.", 13, span=6
    )
    add_tip(ws, "Bronze SHOULD add metadata: _extracted_at, _source_file, _batch_id", 15, span=5)
    add_warning(
        ws, "Bronze should NOT transform, filter, or clean data - that's Silver's job", 17, span=5
    )
    add_tip(
        ws,
        "Bronze = PURE APPEND. Duplicates are expected. Silver handles deduplication.",
        19,
        span=5,
    )

    ws.cell(row=22, column=1, value="Who Fills What")
    ws.cell(row=22, column=1).font = SECTION_FONT
    ws.cell(
        row=23,
        column=1,
        value="• Business stakeholders: B1_Source_Inventory (what data exists, who owns it)",
    )
    ws.cell(row=24, column=1, value="• Engineers: B2-B5 (how to land it technically)")

    ws.cell(row=26, column=1, value="BEFORE YOU START - CHECKLIST:")
    ws.cell(row=26, column=1).font = Font(bold=True, size=11, color="1F4E79")
    checklist = [
        "☐ Do I have access to the source systems?",
        "☐ Do I know who owns this data?",
        "☐ Do I understand what business process creates this data?",
        "☐ Have I talked to someone who uses this data today?",
    ]
    for i, item in enumerate(checklist):
        ws.cell(row=27 + i, column=1, value=item)

    ws.cell(row=32, column=1, value="Odibi Patterns for Bronze")
    ws.cell(row=32, column=1).font = SECTION_FONT

    patterns = [
        ["Pattern", "When to Use", "Key Config"],
        ["Append", "Simple landing, historical accumulation", "mode: append (always)"],
        [
            "Smart Read (Rolling)",
            "Filter by date window automatically",
            "incremental.mode: rolling_window",
        ],
        ["Smart Read (Stateful)", "Track high-water mark", "incremental.mode: stateful"],
    ]
    for i, row_data in enumerate(patterns, 33):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            if i == 33:
                cell.fill = SUBHEADER_FILL
                cell.font = SUBHEADER_FONT

    set_column_widths(ws, [60, 40, 40])

    # B1 - Source Inventory
    ws = wb.create_sheet("B1_Source_Inventory")
    add_title(ws, "SOURCE INVENTORY")
    add_prompt(
        ws,
        "List ALL source systems and tables before designing nodes. One row per source object.",
        2,
    )
    add_encouragement(
        ws,
        "Business folks: This is YOUR sheet. Describe data in your words - engineers translate later.",
        4,
        span=8,
    )

    # Field guidance
    ws.cell(row=6, column=1, value="FIELD GUIDE:")
    ws.cell(row=6, column=1).font = Font(bold=True, size=10)
    field_tips = [
        "Source_System: Where does it come from? (SAP, Salesforce, Excel file, etc.)",
        "Business_Domain: What area of business? (Sales, Production, HR, Finance)",
        "Business_Description: Explain like you're telling a new employee what this data is for",
        "Data_Owner: Who should we call if something looks wrong?",
    ]
    for i, tip in enumerate(field_tips):
        ws.cell(row=7 + i, column=1, value=f"  • {tip}")
        ws.cell(row=7 + i, column=1).font = PROMPT_FONT

    headers = [
        "Source_System",
        "Business_Domain",
        "Source_Type",
        "Source_Identifier",
        "Business_Description",
        "Expected_Daily_Rows",
        "Historical_Range_Needed",
        "Data_Owner",
        "Sensitivity_Level",
        "Contains_PII",
        "Notes",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=12, column=col, value=h)
    style_header_row(ws, 12, len(headers))

    # Prompt row explaining each field
    prompts = [
        "e.g. SAP, Salesforce",
        "Sales, Ops, HR",
        "SQL/API/File",
        "Full path or name",
        "Plain English",
        "~1000, ~10K",
        "1 year, 3 years",
        "Name + role",
        "Public/Internal/Restricted",
        "Yes/No",
        "Any quirks?",
    ]
    for col, p in enumerate(prompts, 1):
        cell = ws.cell(row=13, column=col, value=p)
        cell.font = PROMPT_FONT
        cell.border = THIN_BORDER

    example = [
        "SAP ERP",
        "Production",
        "SQL Table",
        "PROD.dbo.ProductionOrders",
        "Production orders from manufacturing floor",
        "~5,000",
        "3 years",
        "John Smith (Ops)",
        "Internal Only",
        "No",
        "Updates every 15 min",
    ]
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=14, column=col, value=val)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER

    add_data_rows(ws, 15, 40, len(headers))
    set_column_widths(ws, [15, 15, 12, 30, 40, 15, 15, 20, 15, 10, 30])

    # B2 - Node Design (TABLE FORMAT)
    ws = wb.create_sheet("B2_Node_Design")
    add_title(ws, "BRONZE NODE DESIGN")
    add_prompt(
        ws, "One row per Bronze node. All 35+ nodes go here. Fill completely before coding.", 2
    )
    add_tip(ws, "Bronze always uses APPEND mode. Duplicates handled in Silver.", 4, span=12)

    headers = [
        "Node_ID",
        "Node_Name",
        "Source_Ref",
        "Read_Connection",
        "Read_Format",
        "Read_Table_or_Path",
        "Read_Filter",
        "Incremental_Mode",
        "Incremental_Column",
        "Lookback",
        "Lookback_Unit",
        "Fallback_Column",
        "Schema_Policy",
        "Metadata_Columns",
        "Bad_Records_Path",
        "Write_Connection",
        "Write_Path",
        "Tags",
        "Owner",
        "Status",
        "Notes",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(headers))

    # Prompt row
    prompts = [
        "BRZ_001",
        "bronze_xxx",
        "B1 row#",
        "connection name",
        "sql/csv/json",
        "table or path",
        "optional filter",
        "full/rolling/stateful",
        "date column",
        "number",
        "hour/day",
        "backup column",
        "allow/strict",
        "_extracted_at etc",
        "path for bad rows",
        "connection",
        "output path",
        "daily,critical",
        "name",
        "status",
        "",
    ]
    for col, p in enumerate(prompts, 1):
        cell = ws.cell(row=7, column=col, value=p)
        cell.font = PROMPT_FONT
        cell.border = THIN_BORDER

    example = [
        "BRZ_001",
        "bronze_production_orders",
        "B1_Row1",
        "sql_server_prod",
        "sql",
        "PROD.dbo.Orders",
        "",
        "rolling_window",
        "updated_at",
        "3",
        "day",
        "created_at",
        "allow_new_cols",
        "_extracted_at, _batch_id",
        "bronze/bad_records/orders",
        "bronze_lake",
        "bronze/production/orders",
        "daily, critical",
        "Data Team",
        "Planned",
        "Critical for OEE",
    ]
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=8, column=col, value=val)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER

    add_data_rows(ws, 9, 50, len(headers))
    set_column_widths(
        ws, [10, 25, 10, 15, 8, 22, 20, 14, 14, 8, 10, 14, 12, 22, 22, 12, 22, 15, 12, 10, 22]
    )

    inc_mode_dv = DataValidation(
        type="list", formula1='"full_refresh,rolling_window,stateful"', allow_blank=True
    )
    ws.add_data_validation(inc_mode_dv)
    inc_mode_dv.add("H9:H60")

    unit_dv = DataValidation(type="list", formula1='"hour,day,month,year"', allow_blank=True)
    ws.add_data_validation(unit_dv)
    unit_dv.add("K9:K60")

    status_dv = DataValidation(
        type="list", formula1='"Planned,In_Design,Ready_for_Build,In_Prod"', allow_blank=True
    )
    ws.add_data_validation(status_dv)
    status_dv.add("T9:T60")

    # B3 - Column Mapping
    ws = wb.create_sheet("B3_Column_Mapping")
    add_title(ws, "BRONZE COLUMN MAPPING")
    add_prompt(
        ws,
        "Map source columns to Bronze. Bronze = pass-through (no transforms), but document what lands.",
        2,
    )

    headers = [
        "Node_ID",
        "Source_Table",
        "Source_Column",
        "Source_Type",
        "Bronze_Column",
        "Bronze_Type",
        "Is_Key",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    examples = [
        ["BRZ_001", "PROD.dbo.Orders", "order_id", "INT", "order_id", "INT", "Yes", "Primary key"],
        [
            "BRZ_001",
            "PROD.dbo.Orders",
            "order_date",
            "DATETIME",
            "order_date",
            "TIMESTAMP",
            "No",
            "",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=5 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 7, 100, len(headers))
    set_column_widths(ws, [10, 25, 20, 15, 20, 15, 8, 35])

    # B4 - Validation Plan
    ws = wb.create_sheet("B4_Validation")
    add_title(ws, "BRONZE VALIDATION PLAN")
    add_prompt(
        ws,
        "Plan data quality checks BEFORE building. What must be true for this load to be safe?",
        2,
    )
    add_tip(
        ws,
        "Minimum validations: Rowcount vs source, Freshness, Not-null on primary key",
        4,
        span=10,
    )

    headers = [
        "Node_ID",
        "Contract_Type",
        "Target_Column",
        "Plain_English_Goal",
        "Params",
        "On_Failure",
        "Is_Quality_Gate",
        "Owner",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(headers))

    examples = [
        [
            "BRZ_001",
            "not_null",
            "order_id",
            "Primary key must never be null",
            "",
            "error",
            "Yes",
            "Data Team",
            "",
        ],
        [
            "BRZ_001",
            "freshness",
            "_extracted_at",
            "Data should be from today",
            "max_age=24h",
            "warn",
            "No",
            "Data Team",
            "",
        ],
        [
            "BRZ_001",
            "row_count",
            "",
            "Ensure we got rows",
            "min=1",
            "error",
            "Yes",
            "Data Team",
            "",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=7 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 10, 40, len(headers))
    set_column_widths(ws, [10, 15, 15, 35, 20, 10, 12, 12, 30])

    contract_dv = DataValidation(
        type="list",
        formula1='"not_null,unique,accepted_values,row_count,range,regex_match,freshness,schema_contract,custom_sql"',
        allow_blank=True,
    )
    ws.add_data_validation(contract_dv)
    contract_dv.add("B10:B50")

    fail_dv = DataValidation(
        type="list", formula1='"error,warn,filter,quarantine"', allow_blank=True
    )
    ws.add_data_validation(fail_dv)
    fail_dv.add("F10:F50")

    # B5 - Migration
    ws = wb.create_sheet("B5_Migration")
    add_title(ws, "BRONZE MIGRATION MAPPING")
    add_prompt(ws, "For nodes migrating from legacy ETL, document the old approach.", 2)

    headers = [
        "Node_ID",
        "Legacy_Tool",
        "Legacy_Object_Ref",
        "Legacy_Schedule",
        "Legacy_Mode",
        "Known_Pain_Points",
        "Desired_Changes",
        "Cutover_Plan",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))
    add_data_rows(ws, 5, 30, len(headers))
    set_column_widths(ws, [10, 12, 25, 15, 15, 30, 30, 25, 30])

    # Reference
    ws = wb.create_sheet("Reference")
    add_title(ws, "REFERENCE - BRONZE PATTERNS")

    row = 3
    ws.cell(row=row, column=1, value="ANTI-PATTERNS")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="C00000")
    row += 1
    for ap in [
        "Transforming or cleaning data",
        "Filtering business rows",
        "Joining tables",
        "Using Merge pattern",
        "Not adding _extracted_at",
    ]:
        ws.cell(row=row, column=1, value=f"❌ {ap}")
        ws.cell(row=row, column=1).font = Font(color="C00000")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="BEST PRACTICES")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="006100")
    row += 1
    for bp in [
        "Land data as-is",
        "Always APPEND mode",
        "Add _extracted_at metadata",
        "Use Smart Read for incremental",
        "Let Silver deduplicate",
    ]:
        ws.cell(row=row, column=1, value=f"✅ {bp}")
        ws.cell(row=row, column=1).font = Font(color="006100")
        row += 1

    set_column_widths(ws, [80])
    wb.save(output_dir / "Bronze_Walkthrough.xlsx")
    print("Created: Bronze_Walkthrough.xlsx")


# =============================================================================
# SILVER WALKTHROUGH
# =============================================================================


def create_silver_walkthrough(output_dir):
    wb = Workbook()

    # S0 - Overview
    ws = wb.active
    ws.title = "S0_Overview"

    add_title(ws, "SILVER LAYER PLANNING WALKTHROUGH")

    ws.cell(row=3, column=1, value="What is Silver?")
    ws.cell(row=3, column=1).font = SECTION_FONT
    ws.cell(
        row=4,
        column=1,
        value='"Clean, standardize, and enrich ONE source at a time. The best possible version of each source."',
    )

    ws.cell(row=6, column=1, value="The One-Source Test")
    ws.cell(row=6, column=1).font = SECTION_FONT
    ws.cell(row=7, column=1, value='"Could this node run if only ONE source system existed?"')
    ws.cell(row=8, column=1, value="YES → Silver ✓    NO → Probably Gold")
    ws.cell(row=8, column=1).font = Font(bold=True)
    ws.cell(
        row=9,
        column=1,
        value="Note: Reference/lookup table joins ARE allowed (code mappings, enrichment). Only cross-SOURCE-SYSTEM joins go to Gold.",
    )
    ws.cell(row=9, column=1).font = Font(italic=True, size=10)

    steps = [
        "1. Fill S1_Domain_Overview - Business context per domain",
        "2. Fill S2_Node_Design - One row per Silver node (~40 nodes)",
        "3. Fill S3_Column_Mapping - Source → Target with transformations",
        "4. Plan validations in S4_Validation",
    ]
    for i, step in enumerate(steps, 11):
        ws.cell(row=i, column=1, value=step)

    add_encouragement(
        ws, "Silver is where data becomes trustworthy. Take pride in this work.", 16, span=6
    )
    add_tip(
        ws,
        "Silver operations: Deduplicate, Clean text, Cast types, Standardize codes, Enrich via lookups",
        18,
        span=6,
    )
    add_warning(ws, "Do NOT combine multiple source systems here - that's Gold", 20, span=5)
    add_warning(ws, "Do NOT build dimensions/SCD2 here - that's Gold", 22, span=5)

    ws.cell(row=24, column=1, value="Who Fills What")
    ws.cell(row=24, column=1).font = SECTION_FONT
    ws.cell(row=25, column=1, value="• Business: S1_Domain_Overview (what domains exist)")
    ws.cell(row=26, column=1, value="• Engineers: S2-S4 (how to clean and standardize)")

    ws.cell(row=28, column=1, value="BEFORE YOU START - CHECKLIST:")
    ws.cell(row=28, column=1).font = Font(bold=True, size=11, color="1F4E79")
    checklist = [
        "☐ Is the Bronze data landing correctly?",
        "☐ Do I know what 'clean' means for this data?",
        "☐ Have I identified the primary keys for deduplication?",
        "☐ Do I know what codes/abbreviations need to be standardized?",
    ]
    for i, item in enumerate(checklist):
        ws.cell(row=29 + i, column=1, value=item)

    set_column_widths(ws, [70])

    # S1 - Domain Overview
    ws = wb.create_sheet("S1_Domain_Overview")
    add_title(ws, "DOMAIN OVERVIEW")
    add_prompt(ws, "Group your Silver nodes by business domain.", 2)

    headers = [
        "Domain_Name",
        "Business_Description",
        "Key_Bronze_Inputs",
        "Number_of_Nodes",
        "Domain_Owner",
        "Complexity",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))
    add_data_rows(ws, 5, 15, len(headers))
    set_column_widths(ws, [20, 40, 30, 12, 20, 12, 35])

    # S2 - Node Design
    ws = wb.create_sheet("S2_Node_Design")
    add_title(ws, "SILVER NODE DESIGN")
    add_prompt(ws, "One row per Silver node. Fill completely before coding.", 2)
    add_tip(
        ws, "Silver = deduplicate + clean + standardize + enrich. ONE source at a time.", 4, span=12
    )

    headers = [
        "Node_ID",
        "Node_Name",
        "Bronze_Source",
        "Depends_On",
        "Primary_Transformer",
        "Dedup_Keys",
        "Dedup_Order_By",
        "Clean_Text_Cols",
        "Cast_Types",
        "Code_Mappings",
        "Lookup_Joins",
        "Write_Connection",
        "Write_Path",
        "Write_Mode",
        "Tags",
        "Materialization",
        "Owner",
        "Status",
        "Notes",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(headers))

    # Prompt row
    prompts = [
        "SLV_001",
        "silver_xxx",
        "bronze source",
        "other nodes",
        "deduplicate/merge",
        "key columns",
        "order by",
        "columns to clean",
        "col→TYPE",
        "code: old→new",
        "table on key",
        "connection",
        "output path",
        "append/merge",
        "tags",
        "table/view",
        "owner",
        "status",
        "",
    ]
    for col, p in enumerate(prompts, 1):
        cell = ws.cell(row=7, column=col, value=p)
        cell.font = PROMPT_FONT
        cell.border = THIN_BORDER

    example = [
        "SLV_001",
        "silver_orders",
        "brz_orders",
        "",
        "deduplicate",
        "order_id",
        "_extracted_at DESC",
        "customer_name",
        "order_date→TIMESTAMP",
        "status: M1→Machine 1",
        "dim_customer on customer_id",
        "silver_lake",
        "silver/orders",
        "merge",
        "daily",
        "table",
        "Data Team",
        "Planned",
        "",
    ]
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=8, column=col, value=val)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER

    add_data_rows(ws, 9, 50, len(headers))
    set_column_widths(
        ws, [10, 22, 18, 15, 14, 18, 18, 18, 20, 22, 25, 12, 22, 10, 12, 12, 12, 10, 22]
    )

    trans_dv = DataValidation(type="list", formula1='"deduplicate,merge,none"', allow_blank=True)
    ws.add_data_validation(trans_dv)
    trans_dv.add("E9:E60")

    mat_dv = DataValidation(type="list", formula1='"table,view,incremental"', allow_blank=True)
    ws.add_data_validation(mat_dv)
    mat_dv.add("P9:P60")

    # S3 - Column Mapping
    ws = wb.create_sheet("S3_Column_Mapping")
    add_title(ws, "SILVER COLUMN MAPPING")
    add_prompt(ws, "Map Bronze → Silver with transformations. This is your SQL spec.", 2)
    add_encouragement(
        ws, "Document every transformation. Your future self will thank you.", 4, span=8
    )

    headers = [
        "Node_ID",
        "Bronze_Column",
        "Silver_Column",
        "Transformation",
        "Null_Handling",
        "Is_Business_Logic",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(headers))

    examples = [
        [
            "SLV_001",
            "customer_name",
            "customer_name",
            "TRIM(UPPER(x))",
            "COALESCE(x, 'Unknown')",
            "No",
            "Standardize",
        ],
        ["SLV_001", "order_date", "order_date", "TO_TIMESTAMP(x)", "fail if null", "No", "Cast"],
        [
            "SLV_001",
            "machine_code",
            "machine_name",
            "CASE M1→Machine 1",
            "COALESCE(x, 'Unknown')",
            "No",
            "Code mapping",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=7 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_warning(ws, "If Is_Business_Logic = Yes in Silver, consider moving to Gold", 11, span=6)
    add_data_rows(ws, 12, 150, len(headers))
    set_column_widths(ws, [10, 20, 20, 35, 25, 14, 30])

    # S4 - Validation
    ws = wb.create_sheet("S4_Validation")
    add_title(ws, "SILVER VALIDATION PLAN")
    add_prompt(ws, "Silver is the 'trust layer' - plan thorough validation.", 2)

    headers = [
        "Node_ID",
        "Contract_Type",
        "Target_Column",
        "Plain_English_Goal",
        "Params",
        "On_Failure",
        "Is_Quality_Gate",
        "Owner",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    examples = [
        [
            "SLV_001",
            "unique",
            "order_id",
            "No duplicate order_id after dedup",
            "",
            "error",
            "Yes",
            "Data Team",
            "",
        ],
        [
            "SLV_001",
            "not_null",
            "order_id",
            "Business keys must have values",
            "",
            "error",
            "Yes",
            "Data Team",
            "",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=5 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 7, 50, len(headers))
    set_column_widths(ws, [10, 15, 15, 35, 20, 10, 12, 12, 30])

    # Reference
    ws = wb.create_sheet("Reference")
    add_title(ws, "REFERENCE - SILVER PATTERNS")

    row = 3
    ws.cell(row=row, column=1, value="WHAT BELONGS IN SILVER")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="006100")
    row += 1
    for op in [
        "Deduplication",
        "Remove bad characters",
        "Handle nulls",
        "Type casting",
        "Code mapping",
        "Lookup enrichment",
    ]:
        ws.cell(row=row, column=1, value=f"✅ {op}")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="WHAT DOES NOT BELONG")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="C00000")
    row += 1
    for ap in [
        "UNION multiple sources",
        "Business KPIs",
        "Building dimensions with SKs",
        "SCD2 tracking",
        "Cross-fact joins",
    ]:
        ws.cell(row=row, column=1, value=f"❌ {ap}")
        row += 1

    set_column_widths(ws, [80])
    wb.save(output_dir / "Silver_Walkthrough.xlsx")
    print("Created: Silver_Walkthrough.xlsx")


# =============================================================================
# GOLD WALKTHROUGH
# =============================================================================


def create_gold_walkthrough(output_dir):
    wb = Workbook()

    # G0 - Overview
    ws = wb.active
    ws.title = "G0_Overview"

    add_title(ws, "GOLD LAYER PLANNING WALKTHROUGH")

    ws.cell(row=3, column=1, value="What is Gold?")
    ws.cell(row=3, column=1).font = SECTION_FONT
    ws.cell(
        row=4,
        column=1,
        value='"What the business sees. Combine sources, apply business logic, create dimensions, facts, and KPIs."',
    )

    ws.cell(row=6, column=1, value="The Multi-Source Test")
    ws.cell(row=6, column=1).font = SECTION_FONT
    ws.cell(
        row=7,
        column=1,
        value='"Does this require MULTIPLE sources OR business modeling (dimensions/facts)?"',
    )
    ws.cell(row=8, column=1, value="YES → Gold ✓    NO → Probably Silver")
    ws.cell(row=8, column=1).font = Font(bold=True)

    steps = [
        "1. G1_Business_Questions - What does the business need to know?",
        "2. G2_Dimension_Design - Plan dimensions with SCD strategy",
        "3. G3_Fact_Planning - Design fact tables (grain, measures)",
        "4. G4_Node_Design - All Gold nodes (one row per node)",
        "5. G5_Column_Mapping - SQL transformation spec",
        "6. G6_Validation - Reconciliation checks",
        "7. G7_FK_Relationships - Foreign key definitions",
        "8. G8_Semantic_Metrics - Business metrics for self-service",
    ]
    for i, step in enumerate(steps, 10):
        ws.cell(row=i, column=1, value=step)

    add_encouragement(
        ws,
        "Gold is where business value is created. This is what leaders see in dashboards.",
        19,
        span=6,
    )
    add_tip(ws, "Gold = Dimensions + Facts + Aggregations + Business KPIs", 21, span=5)
    add_warning(
        ws, "Gold should NOT do Silver-level cleaning - data should arrive clean", 23, span=5
    )

    ws.cell(row=26, column=1, value="Who Fills What")
    ws.cell(row=26, column=1).font = SECTION_FONT
    ws.cell(
        row=27,
        column=1,
        value="• Business stakeholders: G1_Business_Questions, G8_Semantic_Metrics",
    )
    ws.cell(row=28, column=1, value="• Engineers: G2-G7 (how to model and build)")

    ws.cell(row=30, column=1, value="BEFORE YOU START - CHECKLIST:")
    ws.cell(row=30, column=1).font = Font(bold=True, size=11, color="1F4E79")
    checklist = [
        "☐ Is Silver data clean and deduplicated?",
        "☐ Do I understand what questions the business needs answered?",
        "☐ Have I defined the grain for each fact table?",
        "☐ Do I know which dimensions are needed and their SCD strategy?",
        "☐ Have I talked to report/dashboard consumers about their needs?",
    ]
    for i, item in enumerate(checklist):
        ws.cell(row=31 + i, column=1, value=item)

    set_column_widths(ws, [70])

    # G1 - Business Questions
    ws = wb.create_sheet("G1_Business_Questions")
    add_title(ws, "BUSINESS QUESTIONS & KPIs")
    add_prompt(ws, "Start here! What does the business need to know? Gold answers these.", 2)
    add_encouragement(
        ws,
        "Business folks: Write questions in YOUR words. Engineers will figure out the how.",
        4,
        span=8,
    )

    # Guidance for business users
    ws.cell(row=6, column=1, value="HOW TO FILL THIS:")
    ws.cell(row=6, column=1).font = Font(bold=True, size=10)
    guidance = [
        "Question_or_KPI: What question do you ask in meetings? (e.g., 'What was our OEE yesterday?')",
        "Formula_in_Plain_Words: How would you calculate it by hand? No SQL needed - just words.",
        "Time_Grain: How often do you need this? Daily? Weekly? Monthly?",
        "Breakdowns_Needed: How do you slice it? By plant? By product? By shift?",
        "Criticality: Who looks at this? Executives? Managers? Analysts?",
    ]
    for i, tip in enumerate(guidance):
        ws.cell(row=7 + i, column=1, value=f"  • {tip}")
        ws.cell(row=7 + i, column=1).font = PROMPT_FONT

    add_tip(
        ws,
        "Not sure of the formula? Just describe the DECISION it supports - engineers will help.",
        13,
        span=6,
    )

    headers = [
        "Question_or_KPI",
        "Formula_in_Plain_Words",
        "Time_Grain",
        "Breakdowns_Needed",
        "Gold_Output_Table",
        "Report_Consumers",
        "Criticality",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=15, column=col, value=h)
    style_header_row(ws, 15, len(headers))

    examples = [
        [
            "Daily OEE by line",
            "(Availability × Performance × Quality) × 100",
            "Daily",
            "Line, Plant, Shift",
            "fact_daily_oee",
            "Ops Dashboard",
            "Exec",
        ],
        [
            "Monthly revenue by region",
            "Total order value by month and region",
            "Monthly",
            "Region, Product",
            "agg_monthly_revenue",
            "Finance Report",
            "Regulatory",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=16 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 18, 25, len(headers))
    set_column_widths(ws, [30, 45, 12, 30, 25, 25, 12])

    # G2 - Dimension Design
    ws = wb.create_sheet("G2_Dimension_Design")
    add_title(ws, "DIMENSION DESIGN")
    add_prompt(
        ws, "Design dimensions here. Includes SCD strategy, surrogate keys, history tracking.", 2
    )
    add_warning(ws, "ALWAYS deduplicate Silver data BEFORE applying SCD2!", 4, span=12)

    headers = [
        "Node_ID",
        "Dimension_Name",
        "Silver_Source",
        "Pattern",
        "Natural_Key",
        "Surrogate_Key",
        "SCD_Type",
        "Track_Columns",
        "Effective_From",
        "Effective_To",
        "Is_Current_Col",
        "Unknown_Member",
        "Audit_Columns",
        "Dedup_Before",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(headers))

    examples = [
        [
            "GLD_D01",
            "dim_customer",
            "slv_customers",
            "SCD2",
            "customer_id",
            "customer_sk",
            "Type2",
            "name,email,address",
            "valid_from",
            "valid_to",
            "is_current",
            "Yes",
            "load_ts",
            "Yes",
            "",
        ],
        [
            "GLD_D02",
            "dim_product",
            "slv_products",
            "Dimension",
            "product_id",
            "product_sk",
            "Type1",
            "",
            "",
            "",
            "",
            "Yes",
            "load_ts",
            "Yes",
            "",
        ],
        [
            "GLD_D03",
            "dim_date",
            "N/A",
            "DateDimension",
            "date_id",
            "date_id",
            "Static",
            "",
            "",
            "",
            "",
            "No",
            "",
            "N/A",
            "Generate 2020-2030",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=7 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 10, 20, len(headers))
    set_column_widths(ws, [10, 18, 16, 12, 14, 12, 8, 22, 12, 12, 12, 10, 12, 10, 25])

    pattern_dv = DataValidation(
        type="list", formula1='"Dimension,SCD2,DateDimension"', allow_blank=True
    )
    ws.add_data_validation(pattern_dv)
    pattern_dv.add("D10:D30")

    # G3 - Fact Planning
    ws = wb.create_sheet("G3_Fact_Planning")
    add_title(ws, "FACT TABLE PLANNING")
    add_prompt(ws, "Design each fact table. Grain is critical - define it precisely.", 2)
    add_encouragement(
        ws,
        "If you can say 'One row = one X per Y', you've defined the grain correctly.",
        4,
        span=10,
    )

    headers = [
        "Fact_Name",
        "Business_Question",
        "Grain_Statement",
        "Grain_Columns",
        "Dimension_Lookups",
        "Measures",
        "Pattern",
        "Orphan_Handling",
        "Deduplicate",
        "Audit_Columns",
        "Est_Volume",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(headers))

    example = [
        "fact_daily_prod",
        "G1: Daily OEE",
        "One row = date × location",
        "date_id, location_id",
        "dim_date, dim_location",
        "output_units, downtime_min",
        "Fact",
        "unknown",
        "Yes",
        "load_ts, source=erp",
        "~18K/year",
        "",
    ]
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=7, column=col, value=val)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER

    add_data_rows(ws, 8, 20, len(headers))
    set_column_widths(ws, [20, 18, 30, 22, 30, 30, 12, 12, 10, 18, 14, 25])

    orphan_dv = DataValidation(
        type="list", formula1='"unknown,reject,quarantine"', allow_blank=True
    )
    ws.add_data_validation(orphan_dv)
    orphan_dv.add("H8:H28")

    # G4 - Node Design
    ws = wb.create_sheet("G4_Node_Design")
    add_title(ws, "GOLD NODE DESIGN")
    add_prompt(ws, "One row per Gold node. Fill completely before coding.", 2)

    headers = [
        "Node_ID",
        "Node_Name",
        "Node_Type",
        "Pattern",
        "Design_Sheet_Ref",
        "Depends_On",
        "Silver_Sources",
        "Gold_Sources",
        "Combination_Strategy",
        "Join_Logic",
        "Grain_Statement",
        "Aggregations",
        "KPI_Formulas",
        "Write_Connection",
        "Write_Table",
        "Materialization",
        "Tags",
        "Owner",
        "Status",
        "Notes",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(headers))

    # Prompt row
    prompts = [
        "GLD_001",
        "gold_xxx",
        "type",
        "pattern",
        "G2/G3 ref",
        "upstream",
        "Silver inputs",
        "Gold inputs",
        "union/join",
        "join logic",
        "One row = ?",
        "SUM/AVG/COUNT",
        "formulas",
        "connection",
        "table",
        "table/view",
        "tags",
        "owner",
        "status",
        "",
    ]
    for col, p in enumerate(prompts, 1):
        cell = ws.cell(row=7, column=col, value=p)
        cell.font = PROMPT_FONT
        cell.border = THIN_BORDER

    examples = [
        [
            "GLD_001",
            "combined_prod",
            "Union",
            "union",
            "",
            "slv_prod_a, slv_prod_b",
            "slv_prod_a; slv_prod_b",
            "",
            "union_all",
            "",
            "One row = event",
            "",
            "",
            "gold_lake",
            "gold.combined_prod",
            "table",
            "daily",
            "Data Team",
            "Planned",
            "",
        ],
        [
            "GLD_002",
            "fact_daily_oee",
            "Fact",
            "Fact",
            "G3 Row 7",
            "combined_prod, dim_date",
            "combined_prod",
            "dim_date, dim_location",
            "join",
            "LEFT JOIN on date, location",
            "One row = date × location",
            "SUM(output)",
            "OEE = A×P×Q",
            "gold_lake",
            "gold.fact_oee",
            "table",
            "daily, critical",
            "Data Team",
            "In_Design",
            "",
        ],
        [
            "GLD_D01",
            "dim_customer",
            "Dimension",
            "SCD2",
            "G2 Row 7",
            "slv_customers",
            "slv_customers",
            "",
            "",
            "",
            "One customer version",
            "",
            "",
            "gold_lake",
            "gold.dim_customer",
            "table",
            "daily",
            "Data Team",
            "Planned",
            "",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=8 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 11, 30, len(headers))
    set_column_widths(
        ws, [10, 18, 12, 10, 12, 22, 22, 20, 12, 25, 22, 18, 20, 12, 20, 12, 15, 12, 10, 20]
    )

    type_dv = DataValidation(
        type="list",
        formula1='"Dimension,SCD2,DateDimension,Fact,Aggregation,Union,Join"',
        allow_blank=True,
    )
    ws.add_data_validation(type_dv)
    type_dv.add("C11:C42")

    combo_dv = DataValidation(
        type="list", formula1='"none,union_all,union,join,merge"', allow_blank=True
    )
    ws.add_data_validation(combo_dv)
    combo_dv.add("I11:I42")

    # G5 - Column Mapping
    ws = wb.create_sheet("G5_Column_Mapping")
    add_title(ws, "GOLD COLUMN MAPPING")
    add_prompt(ws, "Map Silver → Gold with transformations. This is your SQL spec.", 2)

    headers = [
        "Node_ID",
        "Source_Table",
        "Source_Column",
        "Target_Column",
        "Transformation",
        "Aggregation",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    examples = [
        [
            "GLD_002",
            "combined_prod",
            "output_qty",
            "total_output",
            "SUM(output_qty)",
            "SUM",
            "Daily aggregate",
        ],
        [
            "GLD_002",
            "calculated",
            "output/target",
            "oee_pct",
            "(output/target)*100",
            "",
            "Business KPI",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=5 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 7, 150, len(headers))
    set_column_widths(ws, [10, 22, 22, 22, 40, 10, 30])

    # G6 - Validation
    ws = wb.create_sheet("G6_Validation")
    add_title(ws, "GOLD VALIDATION & RECONCILIATION")
    add_prompt(ws, "Ensure Gold numbers match expectations. Critical for migrations.", 2)

    headers = [
        "Node_ID",
        "Contract_Type",
        "Business_Description",
        "Params",
        "Reference_Source",
        "On_Failure",
        "Is_Quality_Gate",
        "Owner",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    examples = [
        [
            "GLD_002",
            "custom_sql",
            "Monthly sales = legacy report",
            "query=SELECT SUM...",
            "SSRS_Sales_Report",
            "warn",
            "No",
            "Data Team",
            "Within 1%",
        ],
        [
            "GLD_002",
            "range",
            "OEE between 0-100%",
            "min=0, max=100",
            "",
            "error",
            "Yes",
            "Data Team",
            "",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=5 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 7, 30, len(headers))
    set_column_widths(ws, [10, 15, 35, 30, 25, 10, 12, 12, 30])

    # G7 - FK Relationships (NEW)
    ws = wb.create_sheet("G7_FK_Relationships")
    add_title(ws, "FOREIGN KEY RELATIONSHIPS")
    add_prompt(
        ws, "Define relationships between facts and dimensions for validation and lineage.", 2
    )
    add_tip(ws, "This becomes your relationships.yaml for FK validation", 4, span=8)

    headers = [
        "Relationship_Name",
        "Fact_Table",
        "Dimension_Table",
        "Fact_Key",
        "Dimension_Key",
        "Nullable",
        "On_Violation",
        "Use_In_Lineage",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(headers))

    examples = [
        [
            "orders_to_customers",
            "fact_orders",
            "dim_customer",
            "customer_sk",
            "customer_sk",
            "No",
            "error",
            "Yes",
            "",
        ],
        [
            "orders_to_dates",
            "fact_orders",
            "dim_date",
            "order_date_sk",
            "date_sk",
            "No",
            "error",
            "Yes",
            "",
        ],
        [
            "orders_to_products",
            "fact_orders",
            "dim_product",
            "product_sk",
            "product_sk",
            "Yes",
            "warn",
            "Yes",
            "Nullable for pending orders",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=7 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 10, 30, len(headers))
    set_column_widths(ws, [22, 18, 18, 15, 15, 10, 12, 12, 30])

    viol_dv = DataValidation(
        type="list", formula1='"error,warn,quarantine,filter"', allow_blank=True
    )
    ws.add_data_validation(viol_dv)
    viol_dv.add("G10:G40")

    # G8 - Semantic Metrics (NEW)
    ws = wb.create_sheet("G8_Semantic_Metrics")
    add_title(ws, "SEMANTIC LAYER - METRICS")
    add_prompt(
        ws,
        "Define business metrics for self-service analytics. Non-technical description is key.",
        2,
    )
    add_encouragement(
        ws,
        "Business folks: Define metrics in YOUR words. How would you explain to a new hire?",
        4,
        span=8,
    )

    # Guidance
    ws.cell(row=6, column=1, value="WHY THIS MATTERS:")
    ws.cell(row=6, column=1).font = Font(bold=True, size=10)
    ws.cell(row=7, column=1, value="  • These metrics power dashboards and self-service reports")
    ws.cell(row=7, column=1).font = PROMPT_FONT
    ws.cell(row=8, column=1, value="  • Define them once here, use them everywhere consistently")
    ws.cell(row=8, column=1).font = PROMPT_FONT
    ws.cell(
        row=9,
        column=1,
        value="  • Plain_English_Formula: How would you explain the calculation to someone new?",
    )
    ws.cell(row=9, column=1).font = PROMPT_FONT
    ws.cell(
        row=10,
        column=1,
        value="  • Dimensions_to_Slice: What filters should users be able to apply?",
    )
    ws.cell(row=10, column=1).font = PROMPT_FONT

    add_tip(
        ws,
        "If two reports show different numbers for the same metric, THIS is where you fix it.",
        12,
        span=8,
    )

    headers = [
        "Metric_Name",
        "Business_Question",
        "Plain_English_Formula",
        "SQL_Expression",
        "Source_Fact",
        "Aggregation",
        "Default_Grain",
        "Filters",
        "Dimensions_to_Slice",
        "Owner",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=14, column=col, value=h)
    style_header_row(ws, 14, len(headers))

    examples = [
        [
            "daily_oee",
            "G1: Daily OEE",
            "Availability × Performance × Quality as %",
            "(avail*perf*qual)*100",
            "fact_daily_oee",
            "AVG",
            "day",
            "status='active'",
            "plant, line, shift",
            "Ops Lead",
            "Key exec metric",
        ],
        [
            "total_revenue",
            "G1: Monthly revenue",
            "Sum of all order values",
            "SUM(order_total)",
            "fact_orders",
            "SUM",
            "month",
            "",
            "region, product, customer",
            "Finance",
            "",
        ],
    ]
    for i, ex in enumerate(examples):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=15 + i, column=col, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    add_data_rows(ws, 17, 25, len(headers))
    set_column_widths(ws, [18, 18, 30, 30, 18, 10, 12, 20, 25, 12, 25])

    agg_dv = DataValidation(
        type="list", formula1='"SUM,AVG,COUNT,COUNT_DISTINCT,MIN,MAX"', allow_blank=True
    )
    ws.add_data_validation(agg_dv)
    agg_dv.add("F9:F35")

    # Reference
    ws = wb.create_sheet("Reference")
    add_title(ws, "REFERENCE - GOLD PATTERNS")

    row = 3
    ws.cell(row=row, column=1, value="WHAT BELONGS IN GOLD")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="006100")
    row += 1
    for op in [
        "Dimensions with surrogate keys",
        "SCD2 history tracking",
        "DateDimension generation",
        "Fact tables with SK lookups",
        "UNION multiple sources",
        "Business KPIs",
        "Aggregations",
    ]:
        ws.cell(row=row, column=1, value=f"✅ {op}")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="ANTI-PATTERNS")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="C00000")
    row += 1
    for ap in [
        "Silver-level cleaning",
        "SCD2 on facts",
        "SCD2 without prior dedup",
        "Undefined grain",
    ]:
        ws.cell(row=row, column=1, value=f"❌ {ap}")
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="GRAIN CHECKLIST")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    for gc in [
        "☐ 'One row = one ___' statement defined",
        "☐ Grain columns documented",
        "☐ Duplicate detection validation added",
    ]:
        ws.cell(row=row, column=1, value=gc)
        row += 1

    set_column_widths(ws, [80])
    wb.save(output_dir / "Gold_Walkthrough.xlsx")
    print("Created: Gold_Walkthrough.xlsx")


# =============================================================================
# MAIN
# =============================================================================


def main():
    output_dir = Path(__file__).parent.parent / "templates" / "walkthroughs"
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Creating DE Planning Walkthroughs in: {output_dir}")
    print("=" * 60)

    create_bronze_walkthrough(output_dir)
    create_silver_walkthrough(output_dir)
    create_gold_walkthrough(output_dir)

    print()
    print("=" * 60)
    print("Done! 3 professional-grade walkthrough workbooks created.")
    print()
    print("ENHANCEMENTS IN THIS VERSION:")
    print("  - Validation: Contract types, On_Failure, Is_Quality_Gate")
    print("  - Gold: G7_FK_Relationships for foreign key definitions")
    print("  - Gold: G8_Semantic_Metrics for business metrics layer")
    print("  - All: Tags, Materialization, Depends_On columns")
    print("  - All: Encouragement prompts for non-technical users")
    print("  - Bronze: Lookback_Unit, Fallback_Column for Smart Read")
    print()
    print("By the time you finish filling these out, coding is just typing.")


if __name__ == "__main__":
    main()
