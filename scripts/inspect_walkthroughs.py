"""Inspect walkthrough workbooks for review."""

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")


def inspect_sheet(ws, name):
    print(f"\n=== {name} ===")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")

    # Check first rows for guidance elements
    guidance_found = {
        "title": False,
        "prompt": False,
        "tip": False,
        "warning": False,
        "encourage": False,
        "checklist": False,
        "headers": [],
    }

    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(row, col).value
            if v:
                v_str = str(v)
                if row == 1 and col == 1:
                    guidance_found["title"] = v_str[:50]
                if v_str.startswith("→"):
                    guidance_found["prompt"] = True
                if "✅" in v_str:
                    guidance_found["tip"] = True
                if "⚠️" in v_str:
                    guidance_found["warning"] = True
                if "💡" in v_str:
                    guidance_found["encourage"] = True
                if "☐" in v_str:
                    guidance_found["checklist"] = True

    # Find header row (look for styled row with multiple values)
    for row in range(1, min(20, ws.max_row + 1)):
        row_vals = []
        for col in range(1, min(25, ws.max_column + 1)):
            v = ws.cell(row, col).value
            if v:
                row_vals.append(str(v))
        if len(row_vals) >= 5:  # Likely a header row
            guidance_found["headers"] = row_vals[:10]
            break

    print(f"  Title: {guidance_found['title']}")
    print(f"  Has prompts (→): {guidance_found['prompt']}")
    print(f"  Has tips (✅): {guidance_found['tip']}")
    print(f"  Has warnings (⚠️): {guidance_found['warning']}")
    print(f"  Has encouragement (💡): {guidance_found['encourage']}")
    print(f"  Has checklist (☐): {guidance_found['checklist']}")
    if guidance_found["headers"]:
        print(f"  Headers: {guidance_found['headers']}")


def main():
    base = Path("templates/walkthroughs")

    for xlsx in ["Bronze_Walkthrough.xlsx", "Silver_Walkthrough.xlsx", "Gold_Walkthrough.xlsx"]:
        print(f"\n{'=' * 60}")
        print(f"{xlsx}")
        print("=" * 60)

        wb = load_workbook(base / xlsx)
        for sheet in wb.sheetnames:
            inspect_sheet(wb[sheet], sheet)


if __name__ == "__main__":
    main()
